from datetime import date

from django import forms
from crispy_forms.helper import FormHelper
from crispy_forms.layout import Layout, Submit, Row, Column
from django.db import transaction
from apps.accounts.models import User
from apps.accounts.forms import UserCreateForm
from .models import Student, Teacher, Parent


class StudentForm(forms.ModelForm):
    # User fields embedded in the same form
    first_name   = forms.CharField(max_length=150)
    last_name    = forms.CharField(max_length=150)
    username     = forms.CharField(max_length=150)
    # email        = forms.EmailField()
    password1    = forms.CharField(label='Password', widget=forms.PasswordInput, required=False,
                                   help_text='Leave blank when editing.')
    password2    = forms.CharField(label='Confirm password', widget=forms.PasswordInput, required=False)

    class Meta:
        model  = Student
        fields = ['admission_no', 'date_of_birth', 'gender', 'blood_group', 'student_class', 'address','grades_visible']
        widgets = {'date_of_birth': forms.DateInput(attrs={'type': 'date'})}

    def __init__(self, *args, **kwargs):
        self.user_instance = kwargs.pop('user_instance', None)
        super().__init__(*args, **kwargs)
        if self.user_instance:
            self.fields['first_name'].initial = self.user_instance.first_name
            self.fields['last_name'].initial  = self.user_instance.last_name
            self.fields['username'].initial   = self.user_instance.username
            # self.fields['email'].initial      = self.user_instance.email
        self.helper = FormHelper()
        self.helper.layout = Layout(
            Row(Column('first_name'), Column('last_name')),
            Row(Column('username')),
            Row(Column('admission_no'), Column('date_of_birth')),
            Row(Column('gender'),     Column('blood_group')),
            Row(Column('student_class')),
            'address',
            Row(Column('password1'),  Column('password2')),
            Submit('submit', 'Save student', css_class='btn btn-primary mt-2'),
        )

    def clean(self):
        cleaned = super().clean()
        p1, p2 = cleaned.get('password1'), cleaned.get('password2')
        if p1 and p1 != p2:
            self.add_error('password2', 'Passwords do not match.')
        return cleaned

    def save(self, commit=True):
        student = super().save(commit=False)
        if self.user_instance:
            u = self.user_instance
        else:
            u = User(role='student')
        u.first_name = self.cleaned_data['first_name']
        u.last_name  = self.cleaned_data['last_name']
        u.username   = self.cleaned_data['username']
        # u.email      = self.cleaned_data['email']
        if self.cleaned_data.get('password1'):
            u.set_password(self.cleaned_data['password1'])
        if commit:
            u.save()
            student.user = u
            student.save()
        return student
    


import csv
import io
from openpyxl import load_workbook

class StudentBulkUploadForm(forms.Form):
    file = forms.FileField(
        label='CSV or Excel file',
        help_text='Download the template before uploading.'
    )

    def clean_file(self):
        f = self.cleaned_data['file']
        name = f.name.lower()
        if not name.endswith(('.csv', '.xlsx', '.xls')):
            raise forms.ValidationError('Unsupported file type. Use .csv or .xlsx')
        return f
    
    def _parse_date(self, value):
        """Try multiple date formats and return YYYY-MM-DD string."""
        from datetime import datetime
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        formats = [
            '%Y-%m-%d',   # 2010-10-15
            '%d/%m/%Y',   # 15/10/2010
            '%m/%d/%Y',   # 10/15/2010
            '%d-%m-%Y',   # 15-10-2010
            '%Y/%m/%d',   # 2010/10/15
            '%d %b %Y',   # 15 Oct 2010
            '%d %B %Y',   # 15 October 2010
        ]
        for fmt in formats:
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except ValueError:
                continue
        raise ValueError(f'Unrecognised date format: "{value}". Use DD/MM/YYYY or YYYY-MM-DD.')


    

    def process(self):
        """Call this in the view after is_valid(). Returns list of result dicts."""
        f = self.cleaned_data['file']
        name = f.name.lower()
        rows = []

        if name.endswith('.csv'):
            decoded = f.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        else:
            wb = load_workbook(f)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))

        from apps.academics.models import StudentClass

        # ── Prefetch everything once, instead of querying per row ──
        existing_usernames = set(User.objects.values_list('username', flat=True))
        existing_admission_nos = set(Student.objects.values_list('admission_no', flat=True))
        class_lookup = {c.name.lower(): c for c in StudentClass.objects.all()}

        results = []

        # ── One transaction for the whole batch — huge win on SQLite,
        #    which otherwise fsyncs on every individual save(). ──
        with transaction.atomic():
            for i, row in enumerate(rows, start=2):
                try:
                    first_name   = str(row.get('first_name') or '').strip()
                    last_name    = str(row.get('last_name') or '').strip()
                    username     = str(row.get('username') or '').strip()
                    admission_no = str(row.get('admission_no') or '').strip()
                    raw_dob = str(row.get('date_of_birth') or '').strip()
                    try:
                        dob = self._parse_date(raw_dob)
                    except ValueError as e:
                        results.append({'row': i, 'status': 'error', 'msg': str(e)})
                        continue
                    gender       = str(row.get('gender') or '').strip().upper()
                    password     = str(row.get('password') or 'changeme123').strip()
                    blood_group  = str(row.get('blood_group') or '').strip() or None
                    address      = str(row.get('address') or '').strip()
                    class_name   = str(row.get('class') or '').strip()

                    if not all([first_name, last_name, username, admission_no, dob, gender]):
                        results.append({'row': i, 'status': 'error',
                                        'msg': 'Missing required field(s).'})
                        continue

                    if gender not in ('M', 'F'):
                        results.append({'row': i, 'status': 'error',
                                        'msg': f'Invalid gender "{gender}". Use M or F.'})
                        continue

                    if username in existing_usernames:
                        results.append({'row': i, 'status': 'skip',
                                        'msg': f'Username "{username}" already exists.'})
                        continue

                    if admission_no in existing_admission_nos:
                        results.append({'row': i, 'status': 'skip',
                                        'msg': f'Admission no "{admission_no}" already exists.'})
                        continue

                    student_class = None
                    if class_name:
                        student_class = class_lookup.get(class_name.lower())
                        if not student_class:
                            results.append({'row': i, 'status': 'error',
                                            'msg': f'Class "{class_name}" not found.'})
                            continue

                    u = User(role='student', first_name=first_name,
                            last_name=last_name, username=username)
                    u.set_password(password)
                    u.save()

                    Student.objects.create(
                        user=u,
                        admission_no=admission_no,
                        date_of_birth=dob,
                        gender=gender,
                        blood_group=blood_group,
                        address=address,
                        student_class=student_class,
                    )
                    results.append({'row': i, 'status': 'ok',
                                    'msg': f'{first_name} {last_name} created.'})

                    # Track newly created ones so duplicate rows within
                    # the same file are also caught
                    existing_usernames.add(username)
                    existing_admission_nos.add(admission_no)

                except Exception as e:
                    results.append({'row': i, 'status': 'error', 'msg': str(e)})

        return results

class TeacherForm(forms.ModelForm):
    first_name    = forms.CharField(max_length=150)
    last_name     = forms.CharField(max_length=150)
    username      = forms.CharField(max_length=150)
    email         = forms.EmailField(required=False)
    password1     = forms.CharField(label='Password', widget=forms.PasswordInput, required=False,
                                    help_text='Leave blank when editing.')
    password2     = forms.CharField(label='Confirm password', widget=forms.PasswordInput, required=False)

    class Meta:
        model  = Teacher
        fields = ['staff_id', 'qualification', 'specialization', 'assigned_classes']

    def __init__(self, *args, **kwargs):
        self.user_instance = kwargs.pop('user_instance', None)
        super().__init__(*args, **kwargs)
        if self.user_instance:
            self.fields['first_name'].initial = self.user_instance.first_name
            self.fields['last_name'].initial  = self.user_instance.last_name
            self.fields['username'].initial   = self.user_instance.username
            self.fields['email'].initial      = self.user_instance.email

        self.fields['assigned_classes'].widget = forms.CheckboxSelectMultiple()
        self.fields['assigned_classes'].queryset = self.fields['assigned_classes'].queryset.order_by('name')
            
        self.helper = FormHelper()
        self.helper.layout = Layout(
            Row(Column('first_name'), Column('last_name')),
            Row(Column('username'),   Column('email')),
            Row(Column('staff_id'),   Column('qualification')),
            'specialization',
            'assigned_classes',
            Row(Column('password1'),  Column('password2')),
            Submit('submit', 'Save teacher', css_class='btn btn-primary mt-2'),
        )

    def clean(self):
        cleaned = super().clean()
        p1, p2 = cleaned.get('password1'), cleaned.get('password2')
        if p1 and p1 != p2:
            self.add_error('password2', 'Passwords do not match.')
        return cleaned

    def save(self, commit=True):
        teacher = super().save(commit=False)
        if self.user_instance:
            u = self.user_instance
        else:
            u = User(role='teacher')
        u.first_name = self.cleaned_data['first_name']
        u.last_name  = self.cleaned_data['last_name']
        u.username   = self.cleaned_data['username']
        u.email      = self.cleaned_data['email'] or ''
        if self.cleaned_data.get('password1'):
            u.set_password(self.cleaned_data['password1'])
        if commit:
            u.save()
            teacher.user = u
            teacher.save()
            teacher.assigned_classes.set(self.cleaned_data.get('assigned_classes', []))
        return teacher


class ParentForm(forms.ModelForm):
    first_name = forms.CharField(max_length=150)
    last_name  = forms.CharField(max_length=150)
    username   = forms.CharField(max_length=150)
    email      = forms.EmailField()
    password1  = forms.CharField(label='Password', widget=forms.PasswordInput, required=False,
                                 help_text='Leave blank when editing.')
    password2  = forms.CharField(label='Confirm password', widget=forms.PasswordInput, required=False)
    # ← NO students field here at class level

    class Meta:
        model  = Parent
        fields = ['relation', 'occupation', 'students']  # ← put it back here

    def __init__(self, *args, **kwargs):
        self.user_instance = kwargs.pop('user_instance', None)
        super().__init__(*args, **kwargs)
        if self.user_instance:
            self.fields['first_name'].initial = self.user_instance.first_name
            self.fields['last_name'].initial  = self.user_instance.last_name
            self.fields['username'].initial   = self.user_instance.username
            self.fields['email'].initial      = self.user_instance.email

        self.fields['students'].widget = forms.SelectMultiple(
            attrs={'class': 'select2-students', 'style': 'width:100%'}
        )
        self.fields['students'].label_from_instance = lambda s: f"{s.user.get_full_name()} ({s.admission_no})"
        self.fields['students'].required = False
        self.fields['students'].help_text = 'Search and select one or more students linked to this parent.'

        if self.is_bound:
            # For POST/bound requests, dynamically build the validation queryset from submitted IDs
            student_ids = self.data.getlist('students')
            self.fields['students'].queryset = Student.objects.select_related('user').filter(pk__in=student_ids)
        elif self.instance and self.instance.pk:
            # For GET requests on edit forms, initialize the dropdown with only currently linked students
            self.fields['students'].queryset = self.instance.students.select_related('user').all()
        else:
            # For GET requests on new forms, start with an empty queryset
            self.fields['students'].queryset = Student.objects.none()

        self.helper = FormHelper()
        self.helper.layout = Layout(
            Row(Column('first_name'), Column('last_name')),
            Row(Column('username'),   Column('email')),
            Row(Column('relation'),   Column('occupation')),
            'students',
            Row(Column('password1'),  Column('password2')),
            Submit('submit', 'Save parent', css_class='btn btn-primary mt-2'),
        )

    def clean(self):
        cleaned = super().clean()
        p1, p2 = cleaned.get('password1'), cleaned.get('password2')
        if p1 and p1 != p2:
            self.add_error('password2', 'Passwords do not match.')
        return cleaned

    def save(self, commit=True):
        parent = super().save(commit=False)
        if self.user_instance:
            u = self.user_instance
        else:
            u = User(role='parent')
        u.first_name = self.cleaned_data['first_name']
        u.last_name  = self.cleaned_data['last_name']
        u.username   = self.cleaned_data['username']
        u.email      = self.cleaned_data['email']
        if self.cleaned_data.get('password1'):
            u.set_password(self.cleaned_data['password1'])
        if commit:
            u.save()
            parent.user = u
            parent.save()
            parent.students.set(self.cleaned_data.get('students', []))
        return parent